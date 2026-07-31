# -*- coding: utf-8 -*-
r"""문항 엑셀 '덩어리 틀' — 읽기·마크다운 조립의 순수 규칙 (형태 B, 2026-07-31).

엑셀도 한글도 없이 돈다. VBA 쪽(버튼·덩어리 삽입)은 여기서 못 본다 —
그건 실제 엑셀 COM 으로 실측했다(버튼 보존·매크로 실행 확인).
지키려는 규칙:
  · 빈 값은 `-`(건너뜀) — 값이 한 칸씩 밀리면 시험지가 조용히 어긋난다
  · 여러 줄 값은 { } 덩어리 — 안 묶으면 줄 수만큼 빈칸을 먹는다
  · 시험지 시트가 없으면 None — 예전 표 방식 파일로 넘긴다
"""

import io
import pathlib
import sys
import unittest

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))

from hwp_palette.model import excel_blocks      # noqa: E402


class ToMarkdownTest(unittest.TestCase):
    def test_기본_조립(self):
        blocks = [("가", [("발문", "다음 중 옳은 것은?"), ("선지", ""),
                          ("정답", "3"), ("배점", "4")])]
        md, report, answers = excel_blocks.to_markdown(blocks)
        self.assertEqual("\\가\\\n다음 중 옳은 것은?\n-\n3\n4\n", md)
        self.assertIn("1. 가 — 3/4칸 채움", report[0])
        self.assertEqual([(1, "가", "3", "4")], answers)

    def test_여러_줄은_덩어리로_묶인다(self):
        md, _r, _a = excel_blocks.to_markdown([("가", [("지문", "1줄\n2줄")])])
        self.assertIn("{1줄\n2줄}", md)

    def test_전부_비면_경고가_붙는다(self):
        _md, report, _a = excel_blocks.to_markdown([("가", [("발문", "")])])
        self.assertIn("⚠", report[0])


class ReadBlocksTest(unittest.TestCase):
    def _book(self, sheet_name, rows):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for r, (a, b) in enumerate(rows, start=1):
            if a is not None:
                ws.cell(row=r, column=1, value=a)
            if b is not None:
                ws.cell(row=r, column=2, value=b)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_덩어리를_차례로_읽는다(self):
        buf = self._book("시험지", [
            ("꾸러미 고르기 →", "가"), (None, None),
            ("▶ 가", None), ("발문", "질문"), ("정답", 3), (None, None),
            ("▶ 나", None), ("제목", "중간고사")])
        blocks = excel_blocks.read_blocks(buf)
        self.assertEqual(
            [("가", [("발문", "질문"), ("정답", "3")]),
             ("나", [("제목", "중간고사")])], blocks)

    def test_시험지_시트가_없으면_None(self):
        """예전 표 방식 파일 — 새 판독기가 가로채면 안 된다."""
        buf = self._book("문항", [("번호", "발문")])
        self.assertIsNone(excel_blocks.read_blocks(buf))


if __name__ == "__main__":
    unittest.main()
